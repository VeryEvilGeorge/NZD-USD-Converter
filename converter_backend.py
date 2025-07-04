import openpyxl
import requests
import math
#import customtkinter
#import schedule
#import time

def update_exchange_rate():
    global nzd_to_usd_rate
    api_url = "http://api.exchangeratesapi.io/v1/latest?access_key=19139bb7427f8047329b2c74f9a99e63"  #API url. This one is limited use I think
    try: #scans the database for live exchange rates
        response = requests.get(api_url)
        response.raise_for_status()
        data = response.json()
        nzd_to_usd_rate = data['rates']['NZD']
        return nzd_to_usd_rate
    except Exception as e:
        return print(f"Failed to fetch exchange rate: {e}")

class ExcelConverters:
    def nz_usconverter(columnletter, path, outpath):
        update_exchange_rate() #call latest exchange rate. #It might be worth scheduling this monthly instead.
        columnnum=int(ord((columnletter)))-int(ord('A'))+1 #column number
        #WARNING: This only works for columns A-Z but I never needed that many columns of input data while accounting,
        wb=openpyxl.load_workbook(path) #save the workbook as wb
        sheet=wb['Sheet1'] #sheet name doesn't matter
        for row in range(1,sheet.max_row+1):
            try:
                cell=sheet.cell(row,columnnum) #get cell
                cellval=cell.value #get cell value
                if cellval > 0: #negatives get put to 0.
                    cellval = math.ceil(cellval / nzd_to_usd_rate)  #corrects value, always rounding up for decimals (ceil function)
                else:
                    cellval=0
            except TypeError:
                print(f'No numerical entry in cell C{row}')
            sheet[f'{columnletter}{row}'] = cellval
        wb.save(outpath)
    def us_nzconverter(columnletter, path, outpath):
        update_exchange_rate() #call latest exchange rate. #It might be worth scheduling this monthly instead.
        columnnum=int(ord((columnletter)))-int(ord('A'))+1 #column number
        wb=openpyxl.load_workbook(path) #save the workbook as wb
        sheet=wb['Sheet1'] #sheet name doesn't matter
        for row in range(1,sheet.max_row+1):
            try:
                cell=sheet.cell(row,columnnum) #get cell
                cellval=cell.value #get cell value
                if cellval > 0: #negatives get put to 0.
                    cellval = math.ceil(cellval * nzd_to_usd_rate)  #corrects value, always rounding up for decimals (ceil function)
                else:
                    cellval=0
            except TypeError:
                print(f'No numerical entry in cell C{row}')
            sheet[f'{columnletter}{row}'] = cellval
        wb.save(outpath)
#ExcelConverters.nz_usconverter('C','python_xl_practice.xlsx','python_xl_practice.xlsx')
#Example use: ExcelConverters.nz_usconverter('C','python_xl_practice.xlsx','python_xl_practice.xlsx')
#converts all numbers in column C to their USD equivalent in the file 'python_xl_practice.xlsx'. It is saved to the output file, 'python_xl_practice.xlsx', which is not necessarily the same.